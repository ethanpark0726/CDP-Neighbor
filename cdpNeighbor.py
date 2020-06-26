import datetime
import time
import wexpect
import getpass
import openpyxl
import parse
import json
import os

def createExcelFile():
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CDP Neighbor'
    ws['A2'] = nowDate
    ws['A4'] = 'Hostname'
    ws['B4'] = 'IP Address'
    ws['C4'] = 'Interface'
    ws['D4'] = 'CDP Neighbor'

    wb.save('CDP_Neighbor_Info.xlsx')
    wb.close()

def saveExcelFile(cdpInfo, elem, cellNumber):

    wb = openpyxl.load_workbook('CDP_Neighbor_Info.xlsx')
    ws = wb.active

    ws['A' + str(cellNumber)] = elem[0]
    ws['B' + str(cellNumber)] = elem[2]

    for data in cdpInfo:
        
        ws['C' + str(cellNumber)] = list(data.keys())[0].strip()
        ws['D' + str(cellNumber)] = list(data.values())[0].strip()
        cellNumber += 1

    wb.save('CDP_Neighbor_Info.xlsx')
    wb.close()

def accessJumpBox(username, password):

    print('\n--- Attempting connection to ' + 'IS6 Server... ')
    ssh_newkey = 'Are you sure you want to continue connecting'
    session = wexpect.spawn('ssh ' + username + '@yourserver')

    idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline(password)
    elif idx == 1:
        session.sendline(password)

    idx = session.expect(['$', wexpect.EOF])

    if idx == 0:
        print("--- Successful Login to JumpBox")
        return session
    else:
        print("--- Terminated program")
        exit()

def accessSwitches(session, switch, username, password):

    if 'SSH' in switch[3]:
        ssh_newkey = 'Are you sure you want to continue'
        session.sendline('ssh ' + switch[2])

        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline('yes')
            time.sleep(.5)
            session.sendline(password)
        elif idx == 1:
            session.sendline(password)
        
    else:
        session.sendline('telnet ' + switch[2])
        
        print('\n------------------------------------------------------')
        print('--- Attempting connection to: ' + switch[2])
        print('------------------------------------------------------\n')

        idx = session.expect(['name', wexpect.EOF])

        if idx == 0:
            session.sendline(username)
            idx = session.expect(['word', wexpect.EOF])
            session.sendline(password)

        else:
            print('Something''s wrong!')
            print('--- Terminated Program!!')
            exit()

    idx = session.expect(['>', '#', wexpect.EOF])
    print('--- Success Login to: ', switch[2])
 
    if idx == 0:
        session.sendline('en')
        idx = session.expect(['word:', wexpect.EOF])
        
    if idx == 0:
        session.sendline(password)
        session.expect(['#', wexpect.EOF])
    
    return session

def getDeviceList():
    deviceList = []

    file = open('0624.txt', 'r')

    for line in file:
        temp = line.split('\t')
        temp[-1] = temp[-1].replace('\n', '')
        deviceList.append(temp)
    file.close()

    return deviceList

def commandExecute(session, os):

    command = ''

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])

    if os == 'IOS':
        command += 'sh cdp nei | b Device'
    elif os == 'NXOS':
        command += 'sh cdp nei | b Device-ID'

    session.sendline(command)
    session.expect(['#', wexpect.EOF])

    return session.before.splitlines()

def commandExecuteCDPNeighbor(session, interfaceList, os):
    
    command = ''
    cdpNeighborDump = []
    cdpPair = {}

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])

    if os == 'IOS':
        
        for interface in interfaceList:
            command = 'sh cdp nei ' + interface + ' detail'
            session.sendline(command)
            session.expect(['#', wexpect.EOF])

            for line in session.before.splitlines():
                if 'Device ID' in line:
                    cdpPair[interface] = line
                    cdpNeighborDump.append(cdpPair)
                    cdpPair = {}
                    break
            command = ''
        
    elif os == 'NXOS':
        for interface in interfaceList:
            command = 'sh cdp neighbors interface ' + interface + ' detail'
            session.sendline(command)
            session.expect(['#', wexpect.EOF])
            
            for line in session.before.splitlines():
                if 'Device ID' in line:
                    cdpPair[interface] = line
                    cdpNeighborDump.append(cdpPair)
                    cdpPair = {}
                    break
            
            command = ''

    return cdpNeighborDump

def commandExecutePortChannel(session, listData, baseDesc):
    
    totalDescription = list()

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])

    print(baseDesc)
    for elem in baseDesc:
        
        command = ''
        key = list(elem.keys())[0]
        value = list(elem.values())[0]

        if listData[1] == 'IOS':
            command = command + 'sh etherchannel summary | i ' + key.split(' ')[-1]
        elif listData[1] == 'NXOS':
            command = command + 'sh int ' + key + ' | i Belongs'

        session.sendline(command)
        session.expect(['#', wexpect.EOF])

        portChannel = session.before.splitlines()[1]
        

        if not listData[0].startswith(portChannel.strip()[0:1]):
            
            portChannelNumber = str()

            if listData[1] == 'IOS':

                # Example display
                # 1      Po1(SU)          -        Te1/1/4(P)  Te2/1/4(P)
                temp = portChannel.split('(')[0][::-1]
                
                for char in temp:
                    if char != 'P':
                        portChannelNumber += char
                    else:
                        portChannelNumber += 'P'
                        portChannelNumber = portChannelNumber[::-1]
                        break
                
            elif listData[1] == 'NXOS':
                print(portChannel)
                portChannelNumber = portChannel.split('to')[-1].strip()
                print(portChannelNumber)
            if ('mgmt' not in key) or ('ise' not in value):
                elem[key] = value + ' (' + portChannelNumber + ')'

        totalDescription.append(elem)
    
    print('--- Description assembled Successfully!')
    return totalDescription

if __name__ == '__main__':

    cellNumber = 5
    print()
    print('+-------------------------------------------------------------+')
    print('|    Cisco L2 switches CDP Neighbor Info Gathernig tool...    |')
    print('|    Version 1.0.0                                            |')
    print('|    Compatible with C35xx, C37xx, C38xx, C65XX               |')
    print('|    Nexus 5K, 7K, 9K                                         |')
    print('|    Scripted by Ethan Park, May. 2020                        |')
    print('+-------------------------------------------------------------+')
    print()
    username = input("Enter your admin ID ==> ")
    password = getpass.getpass("Enter your password ==> ")
    print()

    switchList = getDeviceList()
    createExcelFile()

    for elem in switchList:
        
        session = accessJumpBox(username, password)
        result = os.system('ping -n 1 -w 2 ' + elem[2])

        if result == 0:
            session = accessSwitches(session, elem, username, password)
            data = commandExecute(session, elem[1])
            switch = parse.Parse(data, elem[1])
            interfaceList = switch.getInterfaceList(switch.getOS())
            dumpData = commandExecuteCDPNeighbor(session, interfaceList, switch.getOS())
            baseDescription = switch.getBaseDescription(dumpData)
            portChannelData = commandExecutePortChannel(session, elem, baseDescription)
            saveExcelFile(portChannelData, elem, cellNumber)
            cellNumber += len(portChannelData)
        else:
            error = [{'-': 'Needs to check'}]
            saveExcelFile(error, elem, cellNumber)
            cellNumber += 1

        session.close()
